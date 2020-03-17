from PIL import Image
import time
from openpyxl import load_workbook,drawing
from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut
import sys,os


def do_geocode(lat,lon):
    try:
        location = geolocator.reverse(str(lat) + " ," + str(lon) )
        if len(location.address.split(",")) > 6: 
            address = ','.join(location.address.split(",")[0:7])
            return address
    except GeocoderTimedOut:
        print ("Wrong Coordinates. Please use proper coordinates. Example : lat : 37.5281 , lon: -77.8016")
        exit()

def do_rev_geocode(address):
    coo = [0,0]
    try:
        location = geolocator.geocode(address)
        coo [1] = location.longitude 
        coo [1] = "%.4f" % round(coo[1],4)
        coo [0] = location.latitude
        coo [0] = "%.4f" % round(coo[0],4)
        return coo
    except GeocoderTimedOut:
        print ("Wrong address, Please use proper address. Example : 1910 Judes Ferry Rd, Powhatan, VA 23139, USA")


def tmobile(lat,lon):
    try:
        from selenium import webdriver
        from selenium.webdriver.common.keys import Keys
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.wait import WebDriverWait
        from selenium.webdriver.chrome.options import Options
        from PIL import Image
        import time,os
        from openpyxl import load_workbook,drawing
        # coo = 37.5265, -77.8016
        url = 'https://www.t-mobile.com/coverage/coverage-map'
        browser = webdriver.Chrome()
        browser.implicitly_wait(20)
        browser.maximize_window()
        browser.get(url)
        time.sleep(5)
        elem = ""
        browser.switch_to.frame("iframe")
        try: browser.find_element_by_xpath(r'//*[@id="dismiss"]').click()
        except : pass
        #try:
        elem = browser.find_element_by_xpath(r'//*[@id="pccSearchBox"]')
        elem.click()
        elem.send_keys(str(lat) + ", " + str(lon))
        elem.send_keys(Keys.RETURN)
        time.sleep(3)
        browser.find_element_by_xpath(r'//*[@id="mapCanvas"]/div/div/div[8]/div/div/button[2]').click()
        time.sleep(1)
        browser.find_element_by_xpath(r'//*[@id="mapCanvas"]/div/div/div[8]/div/div/button[2]').click()
        time.sleep(4)
        try: browser.find_element_by_xpath(r'//*[@id="dismiss"]').click()
        except : pass
        browser.save_screenshot("image_temp.jpg");
        im = Image.open('image_temp.jpg')
        im = im.crop((5,210, 1910, 630))
        rgb_im = im.convert('RGB')
        rgb_im.save('t-mobile'+'.jpg')
        #os.remove('image_temp.jpg')
        print ("t-mobile.jpg")
        print('=============================================')
        # img = drawing.image.Image('t-mobile'+'.jpg')
        #img.anchor = 'G'+str(x)
        #img = drawing.image.Image('t-mobile'+'.jpg')
        #rgb_im.width = 1800
        #rgb_im.height = 545
        browser.quit()
        return ()
    except Exception as e:
        print ('error on this recodrd. bypassing.')
        print ('error'+ str(e))
        pass
        # =========================  Fished  ==================



def att(address):
    from selenium import webdriver
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.wait import WebDriverWait
    from selenium.webdriver.chrome.options import Options
    from PIL import Image
    import time
    from openpyxl import load_workbook,drawing
    import sys,os
    # coo = 37.5265, -77.8016
    url   = 'https://www.att.com/maps/wireless-coverage.html'

    browser = webdriver.Chrome()
    browser.maximize_window()
    browser.implicitly_wait(20)
    browser.get(url)
    time.sleep(1)
    browser.find_element_by_xpath(r'//*[@id="root"]/div/div/div[2]/div/div/div[2]/a[3]').click()
    elem = ""

    try :
        elem = browser.find_element_by_xpath(r'//*[@id="searchLocation"]')
        elem.click()
        elem.send_keys(address)
        elem.send_keys(Keys.RETURN)
        browser.find_element_by_xpath(r'//*[@id="navPanel"]/a').click()
        time.sleep(1)
        browser.find_element_by_xpath(r'//*[@id="root"]/div/div/div[2]/div/div/div[2]/a[2]').click()
        time.sleep(1)
        browser.find_element_by_xpath(r'//*[@id="root"]/div/div/div[2]/div/div/div[2]/a[2]').click()
        time.sleep(1)
        browser.find_element_by_xpath(r'//*[@id="root"]/div/div/div[2]/div/div/div[2]/a[2]').click()
        time.sleep(12)
        browser.save_screenshot("image_temp.jpg");
        im = Image.open('image_temp.jpg')
        im = im.crop((20, 70, 1860, 825))
        rgb_im = im.convert('RGB')
        rgb_im.save('att'+'.jpg')
        # os.remove('image_temp.jpg')
        browser.quit()
        print('==========================================')
        #### saving information to worksheet
    except :
        pass

def rootmetric(lat,lon):
    from selenium import webdriver
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.support.wait import WebDriverWait
    from selenium.webdriver.chrome.options import Options
    from PIL import Image
    import time
    import sys,os
    xpos = ypos = 0   
    def hex_click (off_x,off_y):
        # move to position with refrence to element :
        try:
            el = browser.find_elements_by_xpath("/html/body/div[3]/div[3]/div[2]/div[1]/div[2]")[0]
            action = webdriver.common.action_chains.ActionChains(browser)
            action.move_to_element_with_offset(el, off_x, off_y)
            action.perform()
            time.sleep(0.1)
            action.click()
            action.perform()
            time.sleep(0.2)
        except:
            pass
            

    def hex_found():
        # find if this hex has signal value:
        try :
            time.sleep(0.5)
            value = browser.find_elements_by_xpath(r'//*[@id="signal_strength_value"]')[0].text
            if (value == '') or str(value) == '– dBm': return(0)
            else: 
                return(value)
        except:
            pass
            return(0)

    def search_hex(x):
        global xpos,ypos
        for r in range(-1,2):
            for c in range(-1,2):
                if (r==0) and (c==0): continue
                hex_click(-20 + (x*r), 200 + (x*c)) 
                xpos = -20 + (x*r)
                ypos = 200 + (x*c)
                if hex_found() != 0:return(hex_found())
        return(0)

    def find_signal():
        global xpos,ypos
        offset = 20
        step  = 15
        s_strength = 0
        while (s_strength == 0):
            try: s_strength = search_hex(offset)
            except : pass
            if s_strength != 0 : break
            offset = offset + step
            if offset > 140 : break

        return(s_strength)

    try:

        browser = webdriver.Chrome()
        browser.implicitly_wait(2)
        browser.maximize_window()
        browser.get("http://webcoveragemap.rootmetrics.com/en-US")  
        time.sleep(3)
        elem = ""
        browser.find_element_by_xpath('/html/body/div[2]/div/div[1]/a').click()
        time.sleep(1)
        try:
            #    ################## AT&T ###############################
            elem = browser.find_element_by_xpath(r'//*[@id="search"]')
            elem.click()
            elem.send_keys(str(lat) + ", " + str(lon))
            elem.send_keys(Keys.RETURN)
            time.sleep(3)
            browser.find_element_by_xpath(r'/html/body/div[3]/div[3]/div[2]/div[1]/div[1]').click()
            time.sleep(1)
            browser.find_element_by_xpath(r'/html/body/div[3]/div[3]/div[2]/div[2]/div[1]/ul/span/li[1]/a').click()
            time.sleep(3)
            browser.find_element_by_xpath(r'/html/body/div[3]/div[3]/div[2]/div[1]/div[1]').click()
            time.sleep(1)
            browser.find_element_by_xpath(r'/html/body/div[5]/div[1]/div/div/div[8]/div/div/button[1]').click()
            browser.find_element_by_xpath(r'/html/body/div[5]/div[1]/div/div/div[8]/div/div/button[1]').click()
            time.sleep(1)
            hex_click(-20, 200)
            s_strength = hex_found()
            if int(s_strength.upper().replace('DBM','').replace(' ','')) == 0: s_strength = find_signal()
            if (s_strength == '') or (str(s_strength) == '- dBm') or (s_strength == 0): s_strength = "Untested area"
            browser.save_screenshot("image_temp.jpg")
            im = Image.open('image_temp.jpg')
            im = im.crop((0, 105, 1900, 730))
            rgb_im = im.convert('RGB')
            rgb_im.save('rm_att.jpg')
            print('provider        :'+ 'AT&T')
            print('Singal Strength :'+ str(s_strength))
            print('Map image file  :'+ 'rm_att'+'.jpg\n')
            
            # #    ################## SPRINT ###############################
            browser.find_element_by_xpath(r'/html/body/div[3]/div[3]/div[2]/div[1]/div[1]').click()
            time.sleep(1)
            browser.find_element_by_xpath(r'/html/body/div[3]/div[3]/div[2]/div[2]/div[1]/ul/span/li[2]/a').click()
            time.sleep(3)
            browser.find_element_by_xpath(r'/html/body/div[3]/div[3]/div[2]/div[1]/div[1]').click()
            time.sleep(1)
            hex_click(-20, 200)
            s_strength = hex_found()
            if s_strength == 0: s_strength = find_signal()
            if (s_strength == '') or (str(s_strength) == '- dBm') or (s_strength == 0): s_strength = "Untested area"
            browser.save_screenshot("image_temp.jpg")
            im = Image.open('image_temp.jpg')
            im = im.crop((0, 105, 1900, 730))
            rgb_im = im.convert('RGB')
            rgb_im.save('rm_sprint.jpg')
            print('provider        :'+ 'Sprint')
            print('Singal Strength :'+ str(s_strength))
            print('Map image file  :'+ 'rm_sprint.jpg\n')
            
            # #    ################## T-Mobile ###############################
            browser.find_element_by_xpath(r'/html/body/div[3]/div[3]/div[2]/div[1]/div[1]').click()
            time.sleep(1)
            browser.find_element_by_xpath(r'/html/body/div[3]/div[3]/div[2]/div[2]/div[1]/ul/span/li[3]/a').click()
            time.sleep(3)
            browser.find_element_by_xpath(r'/html/body/div[3]/div[3]/div[2]/div[1]/div[1]').click()
            time.sleep(1)
            hex_click(-20, 200)
            s_strength = hex_found()
            if s_strength == 0: s_strength = find_signal()
            if (s_strength == '') or (str(s_strength) == '- dBm') or (s_strength == 0): s_strength = "Untested area"
            browser.save_screenshot("image_temp.jpg")
            im = Image.open('image_temp.jpg')
            im = im.crop((0, 105, 1900, 730))
            rgb_im = im.convert('RGB')
            rgb_im.save('rm_t_mobile.jpg')
            print('provider        :'+ 'T-mobile')
            print('Singal Strength :'+ str(s_strength))
            print('Map image file  :'+ 'rm_t_mobile.jpg\n')
            
            #    ################## Verison ###############################
            browser.find_element_by_xpath(r'/html/body/div[3]/div[3]/div[2]/div[1]/div[1]').click()
            time.sleep(1)
            browser.find_element_by_xpath(r'/html/body/div[3]/div[3]/div[2]/div[2]/div[1]/ul/span/li[4]/a').click()
            time.sleep(3)
            browser.find_element_by_xpath(r'/html/body/div[3]/div[3]/div[2]/div[1]/div[1]').click()
            time.sleep(1)
            hex_click(-20, 200)
            s_strength = hex_found()
            if s_strength == 0: s_strength = find_signal()
            if (s_strength == '') or (str(s_strength) == '- dBm') or (s_strength == 0): s_strength = "Untested area"
            browser.save_screenshot("image_temp.jpg")
            im = Image.open('image_temp.jpg')
            im = im.crop((0, 105, 1900, 730))
            rgb_im = im.convert('RGB')
            rgb_im.save('rm_verizon.jpg')
            print('provider        :'+ 'Verizon')
            print('Singal Strength :'+ str(s_strength))
            print('Map image file  :'+ 'rm_verizon.jpg\n')
            browser.quit()
            print('==========================================')
    
        except Exception as e:
            print(e)
            pass

    except Exception as e:
            print(e)
            pass

def sprint(lat,lon):
    from selenium import webdriver
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.wait import WebDriverWait
    from selenium.webdriver.firefox.options import Options
    from PIL import Image
    import time
    from openpyxl import load_workbook,drawing
    import sys

    # coo = 37.5265, -77.8016

    url = 'https://coverage.sprint.com/IMPACT.jsp?'

    from selenium.webdriver.firefox.options import Options

    # options = webdriver.FirefoxOptions()
    # browser = webdriver.Firefox(executable_path=r'geckodriver.exe', options=options)
    extension = 'C:\\Users\\gnjs9314\\AppData\\Roaming\\Mozilla\\Firefox\\Profiles\\j1g2p20i.default\\extensions\\browsec@browsec.com.xpi'
    browser = webdriver.Firefox()
    browser.maximize_window()
    # browser.install_addon(extension,temporary=True)
    browser.implicitly_wait(30)
    # g = input('go ahead ?')
    # print(browser.window_handles)
    browser.switch_to.window(browser.window_handles[0])
    def load_browser():
        browser.get(url)
        time.sleep(5)
        browser.find_element_by_xpath('//*[@id="voicetab"]').click()
        time.sleep(1)
        return()


    load_browser()
    time.sleep(5)
    browser.execute_script("window.scrollTo(0,3000)")
    try:
        elem = ""
        elem = browser.find_element_by_xpath('//*[@id="fulladdress1"]')
        elem.click()
        time.sleep(0.5)
        elem.clear()
        elem.send_keys(str(lat) + ", " + str(lon))
        elem.send_keys(Keys.RETURN)
        time.sleep(4)
        browser.find_element_by_xpath('/html/body/div[2]/div/div[2]/div[1]/div/div/div[8]/div/div/button[2]').click()      # zoom out
        time.sleep(3)
        try:
            #congressional_dist=WebDriverWait(browser, 20).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[2]/div/div[2]/div[1]/div/div/div[1]/div[3]/div/div[4]/div/div/div/div/div/div/h3")))
            elem = browser.find_element_by_xpath('/html/body/div[2]/div/div[2]/div[1]/div/div/div[1]/div[3]/div/div[4]/div/div/div/div/div/div/h3')
            address = elem.text
        except Exception as f:
            print ('Map not loading correctly. Reloading site\n ')
            print ('\n' + str(f))
            browser.find_element_by_xpath('//*[@id="fulladdress1"]').clear()
            load_browser()
            browser.execute_script("window.scrollTo(0,3000)")
            pass

        browser.find_element_by_xpath('/html/body/div[2]/div/div[2]/div[1]/div/div/div[1]/div[3]/div/div[4]/div/div/div/div/button').click() # close address tag
        # elem = browser.find_element_by_xpath('/html/body/div[2]/div/div[2]/div[1]/div/div/button') # maximize map
        elem = browser.find_element_by_css_selector('button.gm-control-active:nth-child(8)')
        elem.click()
        time.sleep(3.5)
        browser.save_screenshot("image_temp.jpg")
        im = Image.open('image_temp.jpg')
        rgb_im = im.convert('RGB')
        rgb_im.save('sprint.jpg')
        elem.click()
        time.sleep(1)
        # elem.send_keys(Keys.Escape)
        print('Map image file:'+ 'sprint.jpg')
        print('==========================================')
        # map container //*[@id="mapcontainer"]
        #### saving information to worksheet
    
    except Exception as f:
        print ("error is: " + str(f))
        pass
    browser.quit()
    # =========================  Fished  ==================

def verizon(address):
    from selenium import webdriver
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.support.wait import WebDriverWait
    from selenium.webdriver.common.action_chains import ActionChains
    from PIL import Image
    import time
    import sys

    def page_has_loaded():
        page_state = browser.execute_script('return document.readyState;')
        return page_state == 'complete'

    # coo = 37.5265, -77.8016
    url   = "https://www.verizonwireless.com/reusable-content/landing-page/coverage-map.html"
    try:
        browser = webdriver.Chrome()
        browser.maximize_window()
        browser.implicitly_wait(20)
        browser.get(url)
        time.sleep(6)
        browser.find_element_by_xpath(r'/html/body/div/div/div/div/section/div/div/div/div/div/div/div/div[1]/div[3]/span[1]/a[1]').click()
        browser.execute_script("window.scrollTo(0,0)")

    except Exception as f:
        print ("error in initial load" + str(f))
        pass

    elem = ""
    try :
        elem = browser.find_element_by_xpath(r'//*[@id="txt-usearch"]')
        elem.clear()
        elem.send_keys(address)
        elem.send_keys(Keys.RETURN)
        time.sleep(4)
        browser.find_element_by_xpath(r'//*[@id="sidebar"]/div[1]/button/span').click()
        time.sleep(5)
        browser.find_element_by_xpath(r'//*[@id="map"]/div[1]/div[3]/div[6]/div/a[2]').click()
        time.sleep(10)
        browser.execute_script("window.scrollTo(0,152)")
        browser.save_screenshot("image_temp.jpg")
        im = Image.open('image_temp.jpg')
        im = im.crop((22, 0, 1873, 848))
        rgb_im = im.convert('RGB')
        rgb_im.save('verizon.jpg')
        # os.remove('image_temp.jpg')
        print('Map image file:'+ 'verizon.jpg')
        print('============================================')
        #### saving information to worksheet

    except Exception as e:
        pass
        print ('error happened ' + str(e))
        print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno), type(e).__name__, e)
        if page_has_loaded() != True :
            browser.refresh()
            time.sleep(6)
            browser.find_element_by_xpath(r'/html/body/div/div/div/div/section/div/div/div/div/div/div/div/div[1]/div[3]/span[1]/a[1]').click()

    browser.quit()   
    # =========================  Fished  ==================

try:
    geolocator = Nominatim(user_agent="GeocodeFarm")
    #coo = [37.5265, -77.8016]
    # lat = 37.5265
    # lon = -77.8016
    # address = "8337 oswego road baldwinsville, ny 13027"
    i = input("Generate map Report by address or coordinates a/c? ")
    if i.upper() == 'A': 
        address = input("enter address ? ")
        coo = do_rev_geocode(address)
        lat = coo[0]; lon = coo[1]
        address = do_geocode(lat,lon)
        print ('Latitude: ' + str(lat)+' Longitude: '+str(lon))
        print ('Detailed address: '+ str(address))
        
    else :
        lat = float(input("enter latitude  :"))
        lon = float(input("enter longitude :"))
        address = do_geocode(lat,lon)
        print ('Address: '+ str(address))


    from threading import Thread
    t1 = Thread (target=tmobile, args=(lat,lon,))
    t2 = Thread (target=att, args=(address,))
    t3 = Thread (target=rootmetric, args=(lat,lon,))
    t4 = Thread (target=sprint, args=(lat,lon,))
    t5 = Thread (target=verizon, args=(address,))
    t1.daemon = t2.daemon = t3.daemon = t4.daemon = t5.daemon = True
    t1.start() ; t2.start()  ; t3.start()  ; t4.start()  ; t5.start() 
    t1.join() ; t2.join() ; t3.join() ; t4.join() ; t5.join() 
    # t4.start() ;  t4.join()
except Exception as e:
    pass
    print ('map error 0 : '+ str(e))

try:   
    # tmobile(lat,lon)
    workbook = load_workbook(filename="map_sheet.xlsx")
    sheet = workbook.active
    sheet.cell(row=2, column=2).value = lat
    sheet.cell(row=2, column=3).value = lon
    sheet.cell(row=2, column=4).value = str(address)
    img = drawing.image.Image('t-mobile'+'.jpg')
    img.width = 1700
    img.height = 520
    img.anchor = 'E2'
    sheet.add_image(img)
except Exception as e:
    pass
    print ("An error happend 1. ERROR :" + str(e))

try:
    # att(address)
    img = drawing.image.Image('att'+'.jpg')
    img.anchor = 'E3'
    img.width = 1050
    img.height = 520
    sheet.add_image(img)
except Exception as e:
    pass
    print ("An error happend 2 . ERROR :" + str(e))

try:
    # rootmetric(lat,lon)
    img = drawing.image.Image('rm_att.jpg')
    img.anchor = 'E5'
    img.width = 1000
    img.height = 520
    sheet.add_image(img)
    img = drawing.image.Image('rm_sprint.jpg')
    img.anchor = 'F5'
    img.width = 1000
    img.height = 520
    sheet.add_image(img)
    img = drawing.image.Image('rm_t_mobile.jpg')
    img.anchor = 'G5'
    img.width = 1000
    img.height = 520
    sheet.add_image(img)
    img = drawing.image.Image('rm_verizon.jpg')
    img.anchor = 'H5'
    img.width = 1000
    img.height = 520
    sheet.add_image(img)
except Exception as e:
    pass
    print ("An error happend 3. ERROR :" + str(e))

try:
    # sprint(lat,lon)
    img = drawing.image.Image('sprint.jpg')
    img.anchor = 'E6'
    img.width = 1000
    img.height = 520
    sheet.add_image(img)
except Exception as e:
    pass
    print ("An error happend 4. ERROR :" + str(e))

try:  
    # verizon(address)
    img = drawing.image.Image('verizon.jpg')
    img.anchor = 'E7'
    img.width = 1050
    img.height = 520
    sheet.add_image(img)

    workbook.save(filename='map_sheet.xlsx')
    workbook.close()
except Exception as e:
    pass
    print ("An error happend 5. ERROR :" + str(e))
    # print ('Error line number :{}'.format(sys.exc_info()[-1].tb_lineno))
    